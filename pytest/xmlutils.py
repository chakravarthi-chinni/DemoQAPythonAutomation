import openpyxl

class Utils:
    def get_rows(file,sheetname):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook.get_sheet_by_name(sheetname)
        return sheet.max_row

    def get_columns(file,sheetname):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook.get_sheet_by_name(sheetname)
        return sheet.max_column

    def xl_read_data(file,sheetname,rowsnum,columnnum):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook.get_sheet_by_name(sheetname)
        return sheet.cell(row=rowsnum,column=columnnum).value

    def xl_write_data(file,sheetname,rowsnum,columnum,data):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook.get_sheet_by_name(sheetname)
        sheet.cell(row=rowsnum,column=columnum).value=data
        workbook.save(file)